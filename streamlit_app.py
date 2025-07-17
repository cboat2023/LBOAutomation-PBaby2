import streamlit as st
import json
import re
from io import BytesIO
import fitz  # PyMuPDF
from google.cloud import vision
from google.oauth2 import service_account
import openai
import openpyxl
import pdfplumber

# Load credentials
openai.api_key = st.secrets["OPENAI"]["OPENAI_API_KEY"]
creds_dict = json.loads(st.secrets["GCP"]["gcp_credentials"])
credentials = service_account.Credentials.from_service_account_info(creds_dict)
vision_client = vision.ImageAnnotatorClient(credentials=credentials)

# Upload CIM file (PDF or image)
st.title("📊 CIM to LBO Model Automation")
uploaded_cim = st.file_uploader("📄 Upload CIM (PDF or Image)", type=["pdf", "png", "jpg", "jpeg"])
uploaded_excel = st.file_uploader("📊 Upload Excel LBO Template", type=["xlsx"])

# Extract text from PDF/image
def extract_text(file):
    file_bytes = file.read()
    if file.type == "application/pdf":
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        is_digital = any(page.get_text() for page in doc)
        if is_digital:
            with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                return "\n".join(page.extract_text() for page in pdf.pages if page.extract_text())
        else:
            text = ""
            for page in doc:
                image_bytes = page.get_pixmap(dpi=300).tobytes("png")
                image = vision.Image(content=image_bytes)
                response = vision_client.document_text_detection(image=image)
                text += response.full_text_annotation.text + "\n"
            return text
    else:
        image = vision.Image(content=file_bytes)
        response = vision_client.document_text_detection(image=image)
        return response.full_text_annotation.text

# Build GPT prompt
def build_ai_prompt(text):
    return f"""
You are analyzing OCR output from a Confidential Information Memorandum (CIM) for an LBO model.

Extract these financials and return ONLY valid JSON:
1. Revenue (3 historical + 6 projected years)
2. EBITDA (group variants under EBITDA_Candidates if multiple found)
3. Maintenance CapEx
4. Acquisition Count

If you find multiple variants of the same metric (e.g., "Adjusted EBITDA", "EBITDA", "Normalized EBITDA"), 
group them under a "_Candidates" key like this:
{{
  "Revenue": [list of revenue values],
  "EBITDA_Candidates": {{
    "Adjusted EBITDA": [values],
    "EBITDA": [values],
    "Normalized EBITDA": [values]
  }},
  "Maintenance_CapEx": [values],
  "Acquisition_Count": number
}}

Text:
{text}
"""

# Helper function to safely parse JSON
def safe_json_parse(text):
    try:
        # Clean up the response text
        cleaned = re.sub(r"```(json)?", "", text).strip()
        return json.loads(cleaned)
    except json.JSONDecodeError as e:
        st.error(f"Failed to parse JSON: {e}")
        st.code(text)
        return None

# Begin processing
if uploaded_cim and uploaded_excel:
    st.success("✅ CIM and Excel uploaded")

    with st.spinner("📖 Extracting text..."):
        raw_text = extract_text(uploaded_cim)

    st.subheader("🔍 Extracted CIM Text")
    with st.expander("View Text"):
        st.text(raw_text)

    prompt = build_ai_prompt(raw_text)
    messages = [
        {"role": "system", "content": "You are a helpful assistant that ONLY outputs valid JSON. Do not include any explanatory text, just the JSON."},
        {"role": "user", "content": prompt}
    ]

    with st.spinner("🤖 Extracting financials with GPT-4..."):
        response = openai.chat.completions.create(
            model="gpt-4",
            messages=messages,
            temperature=0
        )
        response_text = response.choices[0].message.content
        cim_data = safe_json_parse(response_text)

    if cim_data is None:
        st.error("Failed to extract financial data. Please try again.")
        st.stop()

    st.subheader("📈 Extracted Financial Data")
    st.json(cim_data)

    # Step 1: Detect all candidate metrics (e.g., Revenue_Candidates, EBITDA_Candidates)
    candidate_metrics = {}
    for key, val in cim_data.items():
        if key.endswith("_Candidates") and isinstance(val, dict):
            metric_name = key[:-11]  # Remove "_Candidates"
            candidate_metrics[metric_name] = val

    # Initialize session state for selections
    if 'selections_made' not in st.session_state:
        st.session_state.selections_made = False
    if 'user_selections' not in st.session_state:
        st.session_state.user_selections = {}

    # Show metric selection interface
    if candidate_metrics:
        st.subheader("🧠 Multiple Metric Variants Found")
        st.write("Please select the preferred variant for each metric:")
        
        user_selections = {}
        all_selected = True
        
        for metric, options_dict in candidate_metrics.items():
            options = list(options_dict.keys())
            if options:  # Make sure there are options
                selected = st.selectbox(
                    f"Choose variant for {metric}:", 
                    options, 
                    key=f"select_{metric}",
                    index=0
                )
                user_selections[metric] = selected
            else:
                all_selected = False
        
        if all_selected and st.button("✅ Confirm All Selections"):
            # Apply user selections to the data
            for metric, selected_variant in user_selections.items():
                cim_data[metric] = cim_data[f"{metric}_Candidates"][selected_variant]
                # Remove the candidates dict since we've made our selection
                del cim_data[f"{metric}_Candidates"]
            
            st.session_state.selections_made = True
            st.session_state.user_selections = user_selections
            st.success("✅ All metric selections confirmed.")
            st.rerun()  # Refresh the page to continue with the confirmed data
        
        if not st.session_state.selections_made:
            st.warning("Please confirm all metric selections to proceed.")
            st.stop()
    
    # Continue with Excel processing (only if selections are made or no candidates found)
    if st.session_state.selections_made or not candidate_metrics:
        st.subheader("📊 Processing Excel Template...")
        
        # Load Excel and extract metadata
        try:
            workbook = openpyxl.load_workbook(uploaded_excel)
            sheets = {name: workbook[name] for name in workbook.sheetnames}
            metadata = {}
            
            for name, sheet in sheets.items():
                formulas = []
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas.append((cell.coordinate, cell.value))
                metadata[name] = {
                    "max_row": sheet.max_row,
                    "max_col": sheet.max_column,
                    "formulas": formulas[:10]  # Limit to first 10 formulas for brevity
                }

            # Create mapping prompt
            mapping_messages = [
                {"role": "system", "content": "You are a helpful assistant for populating Excel LBO models."},
                {"role": "user", "content": f"Excel metadata: {json.dumps(metadata, indent=2)[:2000]}"},
                {"role": "user", "content": f"CIM extracted financials: {json.dumps(cim_data, indent=2)[:2000]}"},
                {"role": "user", "content": "Map Revenue, EBITDA, CapEx, and Acquisitions to Excel cells. Provide specific cell references and sheet names."}
            ]
            
            with st.spinner("🤖 Determining where to write data in Excel..."):
                mapping_response = openai.chat.completions.create(
                    model="gpt-4",
                    messages=mapping_messages,
                    temperature=0
                )
                gpt_mapping = mapping_response.choices[0].message.content
                
            st.subheader("📌 GPT Mapping Suggestions")
            st.code(gpt_mapping)

            # Allow manual download of the unmodified file for now
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            st.download_button("📥 Download Excel", data=output, file_name="mapped_lbo_model.xlsx")
            
        except Exception as e:
            st.error(f"Error processing Excel file: {e}")

